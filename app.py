import os
from datetime import datetime
import pandas as pd
import streamlit as st
from openpyxl import load_workbook
import plotly.express as px

DATA_FILE = os.path.join(os.path.dirname(__file__), 'data.xlsx')
ASSET_ORDER = [
    'Savings','Core MSCI World','AI & Big Data','Physical Gold','Global Gov Bond',
    'Core EUR Corp Bond','MSCI World Value','MSCI EM','Defence Tech'
]

st.set_page_config(page_title='Portfolio Dashboard', layout='wide')

@st.cache_data(show_spinner=False)
def load_data(_mtime: float):
    xls = pd.ExcelFile(DATA_FILE)
    assets = pd.read_excel(xls, 'Assets')
    tx_pac = pd.read_excel(xls, 'Transactions_PAC')
    tx_manual = pd.read_excel(xls, 'Transactions_Manual')
    month_end = pd.read_excel(xls, 'Month_End')

    # normalize
    for df, cols in [
        (tx_pac, ['Tx_Date','Month']),
        (tx_manual, ['Tx_Date','Month']),
        (month_end, ['Month']),
    ]:
        for c in cols:
            if c in df.columns:
                df[c] = pd.to_datetime(df[c], errors='coerce')

    for c in ['Planned_Amount','Actual_Amount','Flow_Used']:
        if c in tx_pac.columns:
            tx_pac[c] = pd.to_numeric(tx_pac[c], errors='coerce').fillna(0)
    if 'Amount' in tx_manual.columns:
        tx_manual['Amount'] = pd.to_numeric(tx_manual['Amount'], errors='coerce').fillna(0)
    if 'End_Value' in month_end.columns:
        month_end['End_Value'] = pd.to_numeric(month_end['End_Value'], errors='coerce').fillna(0)

    tx_pac['Status'] = tx_pac['Status'].fillna('Auto')
    tx_pac['Flow_Used'] = tx_pac.apply(
        lambda r: 0 if r['Status'] == 'No' else (r['Actual_Amount'] if r['Status'] == 'Edited' and pd.notna(r['Actual_Amount']) and r['Actual_Amount'] != 0 else r['Planned_Amount']),
        axis=1,
    )

    # summary in python
    all_assets = month_end[['Asset']].drop_duplicates().sort_values('Asset')['Asset'].tolist()
    months = sorted(month_end['Month'].dropna().unique())
    rows = []
    prev_map = {}
    for month in months:
        for asset in all_assets:
            end_val = month_end.loc[(month_end['Month'] == month) & (month_end['Asset'] == asset), 'End_Value'].sum()
            flow_pac = tx_pac.loc[(tx_pac['Month'] == month) & (tx_pac['Asset'] == asset), 'Flow_Used'].sum()
            flow_manual = tx_manual.loc[(tx_manual['Month'] == month) & (tx_manual['Asset'] == asset), 'Amount'].sum()
            total_flow = flow_pac + flow_manual
            prev_end = prev_map.get(asset)
            pnl = None if asset == 'Savings' or prev_end is None else end_val - prev_end - total_flow
            perf = None if asset == 'Savings' or prev_end in (None, 0) else pnl / prev_end
            rows.append({
                'Asset': asset, 'Month': month, 'End_Value': end_val,
                'Flow_PAC': flow_pac, 'Flow_Manual': flow_manual, 'Total_Flow': total_flow,
                'Prev_End': prev_end, 'PnL_EUR': pnl, 'Perf_Pct': perf
            })
            prev_map[asset] = end_val
    summary = pd.DataFrame(rows)
    summary['MonthLabel'] = summary['Month'].dt.strftime('%b/%y')
    month_end['MonthLabel'] = month_end['Month'].dt.strftime('%b/%y')
    tx_pac['MonthLabel'] = tx_pac['Month'].dt.strftime('%b/%y')
    tx_manual['MonthLabel'] = tx_manual['Month'].dt.strftime('%b/%y')

    return assets, tx_pac, tx_manual, month_end, summary


def euro(v):
    if pd.isna(v):
        return '-'
    return f"€ {v:,.0f}".replace(',', 'X').replace('.', ',').replace('X', '.')


def pct(v):
    if pd.isna(v):
        return '-'
    return f"{v*100:.1f}%".replace('.', ',')


def save_month_end(edited_df, selected_month):
    wb = load_workbook(DATA_FILE)
    ws = wb['Month_End']
    # header row 1
    # columns: Month Asset Category End_Value
    month_dt = pd.Timestamp(selected_month).to_pydatetime()
    value_map = {r['Asset']: float(r['End_Value']) if pd.notna(r['End_Value']) else 0 for _, r in edited_df.iterrows()}
    for row in range(2, ws.max_row + 1):
        cell_month = ws[f'A{row}'].value
        asset = ws[f'B{row}'].value
        if isinstance(cell_month, datetime) and cell_month.year == month_dt.year and cell_month.month == month_dt.month and asset in value_map:
            ws[f'D{row}'] = value_map[asset]
    wb.save(DATA_FILE)


def save_pac(edited_df, selected_month):
    wb = load_workbook(DATA_FILE)
    ws = wb['Transactions_PAC']
    month_dt = pd.Timestamp(selected_month).to_pydatetime()
    key_map = {}
    for _, r in edited_df.iterrows():
        key_map[r['Asset']] = (r['Status'], None if pd.isna(r['Actual_Amount']) else float(r['Actual_Amount']))
    for row in range(2, ws.max_row + 1):
        cell_month = ws[f'B{row}'].value or ws[f'A{row}'].value
        asset = ws[f'C{row}'].value
        if isinstance(cell_month, datetime) and cell_month.year == month_dt.year and cell_month.month == month_dt.month and asset in key_map:
            status, actual = key_map[asset]
            ws[f'F{row}'] = status
            ws[f'G{row}'] = actual
    wb.save(DATA_FILE)


def append_manual_tx(tx_date, asset, amount, note):
    wb = load_workbook(DATA_FILE)
    ws = wb['Transactions_Manual']
    next_row = ws.max_row + 1
    tx_date = pd.Timestamp(tx_date).to_pydatetime()
    ws[f'A{next_row}'] = tx_date
    ws[f'B{next_row}'] = datetime(tx_date.year, tx_date.month, 1)
    ws[f'C{next_row}'] = asset
    ws[f'D{next_row}'] = 'Manual'
    ws[f'E{next_row}'] = float(amount)
    ws[f'F{next_row}'] = note
    wb.save(DATA_FILE)


mtime = os.path.getmtime(DATA_FILE)
assets, tx_pac, tx_manual, month_end, summary = load_data(mtime)

month_options = sorted(summary['Month'].dropna().unique())
selected_month = st.sidebar.selectbox(
    'Mese', month_options, index=len(month_options)-1,
    format_func=lambda x: pd.Timestamp(x).strftime('%b/%y')
)

st.title('Portfolio Dashboard')

selected = summary[summary['Month'] == selected_month].copy()
selected_non_savings = selected[selected['Asset'] != 'Savings'].copy()
savings_row = selected[selected['Asset'] == 'Savings']

etf_total = selected_non_savings['End_Value'].sum()
savings_total = savings_row['End_Value'].sum()
total_portfolio = etf_total + savings_total
etf_flow = selected_non_savings['Total_Flow'].sum()
etf_pnl = selected_non_savings['PnL_EUR'].fillna(0).sum()

c1, c2, c3, c4, c5 = st.columns(5)
c1.metric('Portfolio totale', euro(total_portfolio))
c2.metric('ETF', euro(etf_total))
c3.metric('Savings', euro(savings_total))
c4.metric('Flow ETF mese', euro(etf_flow))
c5.metric('PnL ETF mese', euro(etf_pnl))

left, right = st.columns([1.1, 1.9])
with left:
    pie_df = pd.DataFrame({'Bucket':['ETF','Savings'], 'Value':[etf_total, savings_total]})
    fig_pie = px.pie(pie_df, names='Bucket', values='Value', hole=0.45)
    fig_pie.update_layout(title='ETF vs Savings')
    st.plotly_chart(fig_pie, use_container_width=True)

    split_df = month_end[(month_end['Month'] == selected_month) & (month_end['Asset'] != 'Savings')].copy()
    split_df = split_df.groupby('Asset', as_index=False)['End_Value'].sum().sort_values('End_Value', ascending=False)
    fig_split = px.pie(split_df, names='Asset', values='End_Value', hole=0.45)
    fig_split.update_layout(title='ETF split')
    st.plotly_chart(fig_split, use_container_width=True)

with right:
    timeline = summary.groupby('Month', as_index=False).agg(
        ETF_Total=('End_Value', lambda s: selected.assign if False else 0)
    )
    timeline = pd.DataFrame({'Month': sorted(summary['Month'].dropna().unique())})
    timeline['ETF_Total'] = timeline['Month'].apply(lambda m: summary[(summary['Month']==m)&(summary['Asset']!='Savings')]['End_Value'].sum())
    timeline['Savings'] = timeline['Month'].apply(lambda m: summary[(summary['Month']==m)&(summary['Asset']=='Savings')]['End_Value'].sum())
    timeline['Total_Flow'] = timeline['Month'].apply(lambda m: summary[(summary['Month']==m)&(summary['Asset']!='Savings')]['Total_Flow'].sum())
    timeline['Total_PnL'] = timeline['Month'].apply(lambda m: summary[(summary['Month']==m)&(summary['Asset']!='Savings')]['PnL_EUR'].fillna(0).sum())
    timeline['MonthLabel'] = timeline['Month'].dt.strftime('%b/%y')

    fig_stack = px.bar(timeline, x='MonthLabel', y=['ETF_Total','Savings'], barmode='stack')
    fig_stack.update_layout(title='Trend portfolio mese per mese', yaxis_title='€')
    st.plotly_chart(fig_stack, use_container_width=True)

    perf_df = selected_non_savings[['Asset','Perf_Pct']].copy().fillna(0)
    fig_perf = px.bar(perf_df, x='Asset', y='Perf_Pct')
    fig_perf.update_layout(title='Performance mensile ETF', yaxis_tickformat='.1%')
    st.plotly_chart(fig_perf, use_container_width=True)

st.markdown('---')

tab1, tab2, tab3 = st.tabs(['Aggiorna month end', 'Controlla PAC del mese', 'Transazioni manuali'])

with tab1:
    st.subheader(f'Valori di fine mese - {pd.Timestamp(selected_month).strftime("%b/%y")}')
    editor_df = month_end[month_end['Month'] == selected_month][['Asset','End_Value']].copy()
    editor_df['Sort'] = editor_df['Asset'].apply(lambda x: ASSET_ORDER.index(x) if x in ASSET_ORDER else 999)
    editor_df = editor_df.sort_values(['Sort','Asset']).drop(columns='Sort')
    edited_month = st.data_editor(editor_df, use_container_width=True, hide_index=True, num_rows='fixed')
    if st.button('Salva month end'):
        save_month_end(edited_month, selected_month)
        st.cache_data.clear()
        st.success('Valori salvati. Ricarica la pagina.')

with tab2:
    st.subheader(f'PAC mensile - {pd.Timestamp(selected_month).strftime("%b/%y")}')
    pac_df = tx_pac[(tx_pac['Month'] == selected_month)][['Asset','Planned_Amount','Status','Actual_Amount','Flow_Used']].copy()
    pac_df = pac_df[pac_df['Asset'] != 'Savings']
    pac_df['Sort'] = pac_df['Asset'].apply(lambda x: ASSET_ORDER.index(x) if x in ASSET_ORDER else 999)
    pac_df = pac_df.sort_values(['Sort','Asset']).drop(columns='Sort')
    edited_pac = st.data_editor(
        pac_df,
        use_container_width=True,
        hide_index=True,
        column_config={
            'Status': st.column_config.SelectboxColumn(options=['Auto','No','Edited'])
        },
        disabled=['Planned_Amount','Flow_Used']
    )
    if st.button('Salva PAC mese'):
        save_pac(edited_pac, selected_month)
        st.cache_data.clear()
        st.success('PAC salvato. Ricarica la pagina.')

with tab3:
    st.subheader('Aggiungi transazione manuale')
    a1, a2, a3, a4 = st.columns([1,1,1,2])
    with a1:
        tx_date = st.date_input('Data', value=pd.Timestamp(selected_month).date())
    with a2:
        asset = st.selectbox('ETF', [a for a in ASSET_ORDER if a != 'Savings'])
    with a3:
        amount = st.number_input('Importo', step=10.0, format='%.2f')
    with a4:
        note = st.text_input('Nota', value='Extra buy / manual edit')
    if st.button('Aggiungi transazione'):
        append_manual_tx(tx_date, asset, amount, note)
        st.cache_data.clear()
        st.success('Transazione aggiunta. Ricarica la pagina.')

    st.markdown('### Storico transazioni manuali')
    st.dataframe(
        tx_manual[['Tx_Date','Asset','Amount','Notes']].sort_values('Tx_Date', ascending=False),
        use_container_width=True,
        hide_index=True,
    )
